import openpyxl


def loadSheet(sheetName, dataOnly):
        return openpyxl.load_workbook(sheetName, data_only=dataOnly)

# Reads all cells out of ExcelSheet
def readCells():
        filePath = 'example.xlsx'
        wb = loadSheet(filePath, True)
        sheet = wb['Tabelle1']

        for i in range (1, sheet.max_column+1):
                for j in range(1, sheet.max_row+1):
                        name = sheet.cell(row=j, column=i).value
                        print(name)

readCells()